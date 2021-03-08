import os
from openpyxl import Workbook, load_workbook

image_path = r"D:\2021tw_pjt\0111.Seadronix\01.RawData\tasked"
save_name = "test.xlsx"

# if not os.path.isfile(save_name):
#     print("make new file as " + save_name)
#     wb = Workbook()
# else:
#     print("load file : " + save_name)
#     wb = load_workbook(save_name)  # create_file.xlsx  파일에서 wb 불러오기

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("image")
wb.remove(wb['Sheet'])


ws["A1"] = "image_path"
ws["B1"] = "properties"
ws["C1"] = "reference_count"

cnt = 2
for (path, dir, files) in os.walk(image_path):
    path = path.replace("\\", "/")
    for filename in files:
        a = str(path).split("/")[-1]
        ws["A" + str(cnt)] = "2021/03/05/" + "{}/{}".format(path[47:], filename)
        ws["B" + str(cnt)] = "seadronix_" + a
        ws["C" + str(cnt)] = 0
        cnt += 1
        # ext = os.path.splitext(filename)[-1]
        # if ext == '.png':
        #     ws["A" + str(idx+2)] = "2021/03/05/" + "{}/{}".format(path[49:], filename)

wb.save(save_name)
wb.close()

# try:
#     a_path = os.listdir(path)
#     b_path = os.listdir(os.path.join(path, a_path[1]))
#     c_path = os.listdir(os.path.join(path, a_path[1], b_path[0]))
#     d_path = os.listdir(os.path.join(path, a_path[1], b_path[0], c_path[0]))
# except:
#     pass


# a_path = os.listdir(path)
# print(a_path)
# print(b_path)
# print(c_path)
# print(d_path)

# a_path = os.listdir(path)
# for a in range(0, len(a_path)):
#     if a_path[0][-4:] == ".png":
#         print(a_path)
#     else:
#         b_path = os.listdir(os.path.join(path, a_path[a]))
#         for b in range(0, len(b_path)):
#             if b_path[0][-4:] == ".png":
#                 print(b_path)
#             else:
#                 c_path = os.listdir(os.path.join(path, a_path[a], b_path[b]))
#                 for c in range(0, len(c_path)):
#                     if c_path[0][-4:] == ".png":
#                         print(c_path)
#                     else:
#                         d_path = os.listdir(os.path.join(path, a_path[a], b_path[b], c_path[c]))
#                         for d in range(0, len(d_path)):
#                             if d_path[0][-4:] == ".png":
#                                 print(d_path)
#                             else:
#                                 print("오류!")

# a_path = os.listdir(path)
# for a in a_path:
#     if a[-4:] != ".png":
#         b_path = os.listdir(os.path.join(path, a))
#         for b in b_path:
#             if b[-4:] != ".png":
#                 c_path = os.listdir(os.path.join(path, a, b))
#                 for c in c_path:
#                     if c[-4:] != ".png":
#                         d_path = os.listdir(os.path.join(path, a, b, c))
#                         print(d_path)
#                     else:
#                         print(c_path)
#             else:
#                 print(b_path)
#     else:
#         print(a_path)
#

# if a_path[0][-4:] == ".png":
#     print("a_path가 이미지 리스트!")
# elif b_path[0][-4:] == ".png":
#     print("b_path가 이미지 리스트!")
# elif c_path[0][-4:] == ".png":
#     print("c_path가 이미지 리스트!")
# elif d_path[0][-4:] == ".png":
#     print("d_path가 이미지 리스트!")
# else:
#     print("오류!")
