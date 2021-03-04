from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["Num", "Eng", "Math"])  # A, B, C
for i in range(1,100):
    ws.append([i, randint(50,100), randint(50, 100)])

col_B = ws["B"]  # eng column 가지고 오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"]  # Eng, Math column 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]  # 1번째 row 가지고 오기
# for cell in row_title:
#     print(cell.value)

row_range = ws[2:6]  # 2번째 ~ 6번째 rows 가지고 오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

row_max_range = ws[2:ws.max_row]
# for rows in row_max_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")  #A/10, AZ/250
#         xy = coordinate_from_string((cell.coordinate))
#         # print(xy, end= " ")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()

# # 전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[2].value)

# # 전체 columns
# print(tuple(ws.columns))
# for columns in tuple(ws.columns):
#     print(columns[0].value)

# for column in ws.iter_cols():  # 전체 columns
#     print(column[0].value)
#
# for rows in ws.iter_rows():  # 전체 rows
#     print(rows[2].value)

# n번째 줄부터 m번째 줄까지, n번째 열부터 m번째 열까지
for row in ws.iter_rows(min_row=1, max_row=10, min_col=2, max_col=3):
     print(row)
#     print(row[0].value, row[1].value)

for col in ws.iter_cols(min_row=1, max_row=3, min_col=1, max_col=4):
    print(col)
    # print(col[0].value, col[1].vlaue)


wb.save("cell_range.xlsx")