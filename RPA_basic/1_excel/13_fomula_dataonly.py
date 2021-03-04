from openpyxl import load_workbook
wb = load_workbook("fomula.xlsx")
ws = wb.active

# 수식 그대로 가져오고 있음
for row in ws.values:
    for cell in row:
        print(cell)

print()

wb = load_workbook("fomula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None으로 출력
for row in ws.values:
    for cell in row:
        print(cell)

