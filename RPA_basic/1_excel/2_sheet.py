from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()  # 새로운 sheet 기본 이름으로 생성
ws.title = "My Sheet"  # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"  # RGB형태로 값을 넣어주면 시트의 탭 색깔 변경

# Sheet, My Sheet, Your Sheet
ws1 = wb.create_sheet("Your Sheet")  # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("New Sheet",2)  # 2번째 위치에 시트 생성

new_ws = wb["New Sheet"]  # Dict 형태로 Sheet에 접근

print(wb.sheetnames)  # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sheet.xlsx")

