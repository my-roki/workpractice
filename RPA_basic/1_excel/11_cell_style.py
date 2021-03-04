from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("cell_range.xlsx")
ws = wb.active

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

a1 = ws["A1"]  # Num
b1 = ws["B1"]  # Eng
c1 = ws["C1"]  # Math

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True)  # 글자색: 빨강, 이탤릭, 굵게
b1.font = Font(color="CC44FF", name="궁서체", strike=True)  # 글자색: 연보라, 글꼴: 궁서체, 취소선
c1.font = Font(color="0000FF", size=20, underline="double")  # 글자색: 파랑, 크기: 20, 밑줄

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 95점 넘는 셀에 대해서 하이라이트 적용
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if cell.column == 1:  # A 번호열은 제외
            continue

        # cell이 정수형 데이터이고 95점 보다 높으면
        if isinstance(cell.value, int) and cell.value > 95:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")  # 배경을 초록색으로 바꿈
            cell.font = Font(color="FF0000")

# 틀 고정
ws.freeze_panes = "B2 "


wb.save("cell_style.xlsx")
