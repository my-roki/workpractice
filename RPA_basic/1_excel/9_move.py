from openpyxl import load_workbook
wb = load_workbook("cell_range.xlsx")
ws = wb.active

# Num Eng Math -> Num Kor Eng Math 으로 변경
ws.move_range("B1:C100", rows=0, cols=1)
ws["B1"].value = "Kor"  # B1셀에 Kor 입력

ws.move_range("C1:C100", rows=2, cols=-1)
wb.save("move_kor.xlsx")
