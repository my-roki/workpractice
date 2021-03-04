from openpyxl import load_workbook
wb1 = load_workbook("cell_range.xlsx")
ws1 = wb1.active

# ws.delete_rows(8)  # 8번째 줄에 있는 7번 학생의 정보 삭제
ws1.delete_rows(8, ws1.max_row)
wb1.save("delete_rows.xlsx")

wb2 = load_workbook("cell_range.xlsx")
ws2 = wb2.active

# ws2.delete_cols(3)  # 3번째 열(C열)에 있는 정보 삭제
ws2.delete_cols(1, ws2.max_column)
wb2.save("delete_cols.xlsx")
