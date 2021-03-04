from openpyxl import load_workbook

wb1 = load_workbook("cell_range.xlsx")
ws1 = wb1.active

# ws.insert_cols(2)  # B번째 열이 비워진다(새로운 빈 열 추가)
ws1.insert_cols(2, 4)  # B번째 열로부터 4열 추가
wb1.save("insert_cols.xlsx")


wb2 = load_workbook("cell_range.xlsx")
ws2 = wb2.active

# ws2.insert_rows(10)  # 10번째 줄이 비워진다(새로운 빈 행 추가)
ws2.insert_rows(10, 4)  # 10번째 줄로부터 4열 추가
wb2.save("insert_rows.xlsx")
