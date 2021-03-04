from openpyxl import load_workbook
wb = load_workbook("cell_range.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # Num, Eng, Math
    if int(row[1].value) > 95:
        print(row[0].value, "번 학생은 영어 천재")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "Eng":
            cell.value = "Kor"

wb.save("cell_range_modified.xlsx")