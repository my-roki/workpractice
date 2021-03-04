from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 기본 틀 만들기
list = ["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"]
# print(len(list))

scores = [
    [1,10,8,5,14,26,12],
    [2,7,3,7,15,24,18],
    [3,9,5,8,8,12,4],
    [4,7,8,7,17,21,18],
    [5,7,8,7,16,25,15],
    [6,3,5,8,8,17,0],
    [7,4,9,10,16,27,18],
    [8,6,6,6,15,19,17],
    [9,10,10,9,19,30,19],
    [10,9,8,8,20,25,20]
]
# print(len(a))
# print(a[0][1])
# for i in range(2, len(a)+2):
#     print(i)

for i in range(0, len(list)):
    ws.cell(row=1, column=i+1, value=list[i])
    for j in range(2, len(scores)+2):
        ws.cell(row=j, column=i+1, value=scores[j-2][i])

# 퀴즈2 올 만점으로 변경
for k in range(2, len(scores)+2):
    ws.cell(row=k, column=4, value=10)

# H열에 총점 생성 후 총점 계산, I열에 성적정보 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)

    grade = None
    if score[1] <= 5:
        grade = "F"
    elif sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

    ws.cell(row=idx, column=9, value=grade)


wb.save("scores.xlsx")
